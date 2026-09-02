"""Unit tests for auction XLSX export workbook builder."""
from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from advisor.auction_export import build_workbook, sanitize_sheet_name


PLAYERS = [
    {"id": 1, "nome": "Portiere Uno", "ruolo": "P", "squadra": "MIL"},
    {"id": 2, "nome": "Difensore Due", "ruolo": "D", "squadra": "INT"},
    {"id": 3, "nome": "Centrocampista Tre", "ruolo": "C", "squadra": "JUV"},
    {"id": 4, "nome": "Attaccante Quattro", "ruolo": "A", "squadra": "ROM"},
]

RULES = {
    "starting_credits": 750,
    "auction": {"role_budget_percentages": {"P": 7, "D": 18, "C": 25, "A": 50}},
}


class AuctionExportTests(unittest.TestCase):
    def test_build_workbook_basic(self):
        payload = {
            "teams": [
                {"name": "Alpha", "starting_credits": 750},
                {"name": "Beta", "starting_credits": 500},
            ],
            "history": [
                {"player_id": 1, "owner": 0, "price": 20},
                {"player_id": 2, "owner": 0, "price": 40},
                {"player_id": 3, "owner": 0, "price": 60},
                {"player_id": 4, "owner": 0, "price": 100},
                {"player_id": 4, "owner": 1, "price": 50},
            ],
            "role_budget_percentages": {"P": 7, "D": 18, "C": 25, "A": 50},
        }
        workbook = load_workbook(BytesIO(build_workbook(payload, PLAYERS, RULES)))
        self.assertEqual(workbook.sheetnames, ["Alpha", "Beta"])

        alpha = workbook["Alpha"]
        self.assertEqual(alpha["A1"].value, "PORTIERI")
        self.assertEqual(alpha["C1"].value, "DIFENSORI")
        self.assertEqual(alpha["E1"].value, "CENTROCAMPISTI")
        self.assertEqual(alpha["G1"].value, "ATTACCANTI")
        self.assertEqual(alpha["A2"].value, "Nome")
        self.assertEqual(alpha["B2"].value, "Prezzo")
        self.assertEqual(alpha["A3"].value, "Portiere Uno")
        self.assertEqual(alpha["B3"].value, 20)

        # Role saldo = budget - spent for P: round(750*7/100)=52 → 52-20=32
        self.assertEqual(alpha["A6"].value, "Saldo ruolo")
        self.assertEqual(alpha["B6"].value, 32)

        # Overall: 750 - (20+40+60+100) = 530
        self.assertIn("530", str(alpha["A8"].value))

    def test_build_workbook_empty_history(self):
        payload = {
            "teams": [{"name": "Vuota", "starting_credits": 750}],
            "history": [],
            "role_budget_percentages": {"P": 7, "D": 18, "C": 25, "A": 50},
        }
        workbook = load_workbook(BytesIO(build_workbook(payload, PLAYERS, RULES)))
        sheet = workbook["Vuota"]
        # No picks → first data row is Speso ruolo at row 3
        self.assertEqual(sheet["A3"].value, "Speso ruolo")
        self.assertEqual(sheet["B3"].value, 0)
        self.assertEqual(sheet["A5"].value, "Saldo ruolo")
        self.assertEqual(sheet["B5"].value, 52)  # full P budget
        self.assertIn("750", str(sheet["A7"].value))

    def test_build_workbook_unknown_player(self):
        payload = {
            "teams": [{"name": "Solo", "starting_credits": 750}],
            "history": [
                {"player_id": 999, "owner": 0, "price": 99},
                {"player_id": 1, "owner": 0, "price": 10},
            ],
        }
        workbook = load_workbook(BytesIO(build_workbook(payload, PLAYERS, RULES)))
        sheet = workbook["Solo"]
        self.assertEqual(sheet["A3"].value, "Portiere Uno")
        self.assertEqual(sheet["B3"].value, 10)
        # Unknown skipped: overall 750-10
        self.assertIn("740", str(sheet["A8"].value))

    def test_build_workbook_sheet_name_sanitization(self):
        payload = {
            "teams": [
                {"name": "Foo[Bar]:*?/\\Baz", "starting_credits": 750},
                {"name": "Foo_Bar_____Baz", "starting_credits": 750},
            ],
            "history": [],
        }
        workbook = load_workbook(BytesIO(build_workbook(payload, PLAYERS, RULES)))
        for title in workbook.sheetnames:
            self.assertLessEqual(len(title), 31)
            for char in "[]:*?/\\":
                self.assertNotIn(char, title)
        self.assertEqual(len(set(workbook.sheetnames)), 2)

    def test_sanitize_sheet_name_uniqueness(self):
        used: set[str] = set()
        first = sanitize_sheet_name("Dup", used)
        second = sanitize_sheet_name("Dup", used)
        self.assertEqual(first, "Dup")
        self.assertEqual(second, "Dup_2")

    def test_camel_case_history_keys(self):
        payload = {
            "teams": [{"name": "Camel", "startingCredits": 500}],
            "history": [{"playerId": 1, "owner": 0, "price": 15}],
        }
        workbook = load_workbook(BytesIO(build_workbook(payload, PLAYERS, {})))
        sheet = workbook["Camel"]
        self.assertEqual(sheet["A3"].value, "Portiere Uno")
        self.assertEqual(sheet["B3"].value, 15)


if __name__ == "__main__":
    unittest.main()
