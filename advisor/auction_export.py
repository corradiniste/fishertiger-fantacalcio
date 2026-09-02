"""Build an XLSX workbook summarizing live-auction picks per league team."""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROLE_ORDER = ("P", "D", "C", "A")
ROLE_HEADERS = {
    "P": "PORTIERI",
    "D": "DIFENSORI",
    "C": "CENTROCAMPISTI",
    "A": "ATTACCANTI",
}
DEFAULT_ROLE_BUDGET_PERCENTAGES = {"P": 7, "D": 18, "C": 25, "A": 50}
_ILLEGAL_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_YELLOW = PatternFill("solid", fgColor="FFF2CC")
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_HEADER_FILL = PatternFill("solid", fgColor="1F3A2A")
_HEADER_FONT = Font(bold=True, color="F4FAEE")
_SUB_FONT = Font(bold=True)
_BOLD = Font(bold=True)
_TOP = Alignment(vertical="top", wrap_text=True, horizontal="left")
_CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _as_dict(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _pick(mapping: dict[str, Any], *keys: str, default: Any = None) -> Any:
    for key in keys:
        if key in mapping and mapping[key] is not None:
            return mapping[key]
    return default


def role_budget_percentages(rules: dict[str, Any] | None = None, payload: dict[str, Any] | None = None) -> dict[str, float]:
    """Resolve P/D/C/A budget shares from payload, rules, or defaults."""
    payload = _as_dict(payload)
    rules = _as_dict(rules)
    auction = _as_dict(rules.get("auction"))
    raw = _pick(
        payload,
        "role_budget_percentages",
        "roleBudgetPercentages",
        default=None,
    )
    if not isinstance(raw, dict):
        raw = _pick(auction, "role_budget_percentages", "roleBudgetPercentages", default=None)
    if not isinstance(raw, dict):
        raw = _pick(rules, "role_budget_percentages", "roleBudgetPercentages", default=None)
    if not isinstance(raw, dict):
        raw = {}
    resolved: dict[str, float] = {}
    for role in ROLE_ORDER:
        try:
            value = float(raw.get(role, DEFAULT_ROLE_BUDGET_PERCENTAGES[role]))
        except (TypeError, ValueError):
            value = float(DEFAULT_ROLE_BUDGET_PERCENTAGES[role])
        if value < 0:
            value = float(DEFAULT_ROLE_BUDGET_PERCENTAGES[role])
        resolved[role] = value
    return resolved


def sanitize_sheet_name(name: str, used: set[str] | None = None) -> str:
    """Excel sheet titles: max 31 chars, no []:*?/\\, unique within workbook."""
    used = used if used is not None else set()
    cleaned = _ILLEGAL_SHEET_CHARS.sub("_", str(name or "").strip()) or "Squadra"
    cleaned = cleaned[:31]
    base = cleaned
    index = 2
    occupied = {item.casefold() for item in used}
    while cleaned.casefold() in occupied:
        suffix = f"_{index}"
        cleaned = f"{base[: max(0, 31 - len(suffix))]}{suffix}"
        index += 1
    used.add(cleaned)
    return cleaned


def _history_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    history = payload.get("history")
    if not isinstance(history, list):
        return []
    items: list[dict[str, Any]] = []
    for entry in history:
        if not isinstance(entry, dict):
            continue
        player_id = _pick(entry, "player_id", "playerId")
        owner = entry.get("owner")
        price = entry.get("price")
        try:
            owner_index = int(owner)
            price_value = float(price)
            player_key = int(player_id) if str(player_id).lstrip("-").isdigit() else player_id
        except (TypeError, ValueError):
            continue
        items.append({"player_id": player_key, "owner": owner_index, "price": price_value})
    return items


def _teams(payload: dict[str, Any], default_starting: int) -> list[dict[str, Any]]:
    teams = payload.get("teams")
    if not isinstance(teams, list):
        return []
    resolved: list[dict[str, Any]] = []
    for team in teams:
        if not isinstance(team, dict):
            continue
        name = str(_pick(team, "name", default="") or "Squadra")
        starting = _pick(team, "starting_credits", "startingCredits", default=default_starting)
        try:
            starting_credits = int(starting)
        except (TypeError, ValueError):
            starting_credits = default_starting
        resolved.append({"name": name, "starting_credits": starting_credits})
    return resolved


def _default_starting_credits(rules: dict[str, Any], payload: dict[str, Any]) -> int:
    credits = _as_dict(rules.get("credits"))
    candidate = _pick(
        payload,
        "starting_credits",
        "startingCredits",
        default=_pick(
            rules,
            "starting_credits",
            "startingCredits",
            default=_pick(credits, "starting", default=750),
        ),
    )
    try:
        return int(candidate)
    except (TypeError, ValueError):
        return 750


def _players_by_id(players: list[dict[str, Any]]) -> dict[Any, dict[str, Any]]:
    index: dict[Any, dict[str, Any]] = {}
    for player in players:
        if not isinstance(player, dict):
            continue
        player_id = player.get("id")
        if player_id is None:
            continue
        index[player_id] = player
        try:
            index[int(player_id)] = player
        except (TypeError, ValueError):
            pass
        index[str(player_id)] = player
    return index


def _write_role_block(
    ws,
    *,
    col_name: int,
    col_price: int,
    picks: list[tuple[str, float]],
    starting_credits: int,
    budget_percent: float,
) -> int:
    """Write picks + role totals; return last used row index."""
    row = 3
    for nome, price in picks:
        ws.cell(row=row, column=col_name, value=nome).alignment = _TOP
        price_cell = ws.cell(row=row, column=col_price, value=price)
        price_cell.alignment = _TOP
        price_cell.number_format = "0"
        row += 1
    spent = sum(price for _, price in picks)
    budget = round(starting_credits * budget_percent / 100)
    saldo = budget - spent

    spent_label = ws.cell(row=row, column=col_name, value="Speso ruolo")
    spent_label.font = _BOLD
    spent_value = ws.cell(row=row, column=col_price, value=spent)
    spent_value.font = _BOLD
    spent_value.number_format = "0"
    row += 1

    budget_label = ws.cell(row=row, column=col_name, value="Budget ruolo")
    budget_label.font = _BOLD
    budget_value = ws.cell(row=row, column=col_price, value=budget)
    budget_value.font = _BOLD
    budget_value.number_format = "0"
    row += 1

    saldo_label = ws.cell(row=row, column=col_name, value="Saldo ruolo")
    saldo_label.font = _BOLD
    saldo_label.fill = _YELLOW
    saldo_value = ws.cell(row=row, column=col_price, value=saldo)
    saldo_value.font = _BOLD
    saldo_value.fill = _YELLOW
    saldo_value.number_format = "0"
    return row


def build_workbook(
    payload: dict[str, Any],
    players: list[dict[str, Any]],
    rules: dict[str, Any] | None = None,
) -> bytes:
    """Return XLSX bytes: one sheet per team, four role column pairs + balances."""
    payload = _as_dict(payload)
    rules = _as_dict(rules)
    budgets = role_budget_percentages(rules, payload)
    default_starting = _default_starting_credits(rules, payload)
    teams = _teams(payload, default_starting)
    if not teams:
        raise ValueError("export payload requires a non-empty teams array")
    history = _history_items(payload)
    by_id = _players_by_id(players if isinstance(players, list) else [])

    wb = Workbook()
    # Remove the default sheet after the first real one is created.
    default_sheet = wb.active
    used_names: set[str] = set()

    for team_index, team in enumerate(teams):
        sheet_title = sanitize_sheet_name(team["name"], used_names)
        if team_index == 0:
            ws = default_sheet
            ws.title = sheet_title
        else:
            ws = wb.create_sheet(title=sheet_title)

        for role_index, role in enumerate(ROLE_ORDER):
            col_name = role_index * 2 + 1
            col_price = col_name + 1
            start = get_column_letter(col_name)
            end = get_column_letter(col_price)
            ws.merge_cells(f"{start}1:{end}1")
            header = ws.cell(row=1, column=col_name, value=ROLE_HEADERS[role])
            header.font = _HEADER_FONT
            header.fill = _HEADER_FILL
            header.alignment = _CENTER
            name_header = ws.cell(row=2, column=col_name, value="Nome")
            price_header = ws.cell(row=2, column=col_price, value="Prezzo")
            name_header.font = _SUB_FONT
            price_header.font = _SUB_FONT
            name_header.alignment = _CENTER
            price_header.alignment = _CENTER
            ws.column_dimensions[start].width = 32
            ws.column_dimensions[end].width = 10

        team_picks = [item for item in history if item["owner"] == team_index]
        picks_by_role: dict[str, list[tuple[str, float]]] = {role: [] for role in ROLE_ORDER}
        total_spent = 0.0
        for item in team_picks:
            player = by_id.get(item["player_id"])
            if player is None:
                continue
            role = str(player.get("ruolo") or "").upper()
            if role not in picks_by_role:
                continue
            nome = str(player.get("nome") or f"#{item['player_id']}")
            picks_by_role[role].append((nome, item["price"]))
            total_spent += item["price"]

        for role in ROLE_ORDER:
            picks_by_role[role].sort(key=lambda pair: (-pair[1], pair[0].casefold()))

        last_rows: list[int] = []
        for role_index, role in enumerate(ROLE_ORDER):
            col_name = role_index * 2 + 1
            col_price = col_name + 1
            last_rows.append(
                _write_role_block(
                    ws,
                    col_name=col_name,
                    col_price=col_price,
                    picks=picks_by_role[role],
                    starting_credits=team["starting_credits"],
                    budget_percent=budgets[role],
                )
            )

        summary_row = max(last_rows) + 2
        ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)
        summary = ws.cell(
            row=summary_row,
            column=1,
            value=f"Saldo complessivo: {team['starting_credits'] - total_spent:g}",
        )
        summary.font = _BOLD
        summary.fill = _GREEN
        summary.alignment = _CENTER
        for col in range(1, 9):
            ws.cell(row=summary_row, column=col).fill = _GREEN

        ws.freeze_panes = "A3"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
